# -*- coding: utf-8 -*-
"""Pack comité xlsx — généré depuis le data model du cockpit (data.js → data_dump.json).
Agrégats, scénario Bear et contrôles en FORMULES Excel (auditables), inputs en bleu.
Usage : node -e "require('fs').writeFileSync('tools/data_dump.json', JSON.stringify(require('./data.js')))"
        python tools/build_pack_xlsx.py
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = json.load(open(os.path.join(HERE, "data_dump.json"), encoding="utf-8"))
OUT = os.path.join(HERE, "..", "pack", "pack_comite_core_plus_france.xlsx")

NAVY, GOLD, IVORY, LINE = "14213D", "C9A24B", "F7F2EA", "D9D2C2"
BLUE_IN, GREEN_LINK = "0000FF", "008000"
F = lambda **k: Font(name="Arial", size=10, **k)
HDR = dict(font=Font(name="Arial", size=10, bold=True, color="FFFFFF"),
           fill=PatternFill("solid", start_color=NAVY),
           alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
THIN = Border(*[Side(style="thin", color=LINE)] * 4)
M2, PCT1, PCT2, X2 = '#,##0.00;(#,##0.00)', '0.0%', '0.00%', '0.00"x"'

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill, cell.alignment = HDR["font"], HDR["fill"], HDR["alignment"]

def put(ws, ref, value, fmt=None, color=None, bold=False, italic=False, border=True):
    cell = ws[ref]
    cell.value = value
    cell.font = F(color=color or "000000", bold=bold, italic=italic)
    if fmt: cell.number_format = fmt
    if border: cell.border = THIN
    return cell

wb = Workbook()

# ============ Onglet 1 — Lisez-moi ============
ws = wb.active
ws.title = "Lisez-moi"
ws.sheet_properties.tabColor = NAVY
ws.column_dimensions["A"].width = 110
rows = [
    ("PACK COMITÉ — FONDS CORE+ FRANCE (FICTIF)", dict(bold=True, size=14)),
    ("Situation au 31/03/2026 · généré automatiquement depuis le data model du cockpit (data.js)", dict(italic=True)),
    ("", {}),
    ("Ce classeur reproduit, en formules Excel auditables, le moteur de calcul du cockpit web :", {}),
    ("  · Grille actifs — inputs par actif ; LTV et contrôle NOI ÷ cap en formules", {}),
    ("  · Dette & couverture — exposition nette, service, écarts vs BP en formules", {}),
    ("  · Agrégats & covenants — GAV, NAV, VL, DSCR, ratios : aucune valeur saisie, tout est dérivé", {}),
    ("  · Scénario Bear — hypothèses paramétrables (loyers −10 %, vacance +400 bps, caps +125 bps, taux +150 bps)", {}),
    ("  · Contrôles — recette PASS / FAIL comparant le classeur aux valeurs de référence du cockpit", {}),
    ("", {}),
    ("Code couleur : BLEU = input modifiable · NOIR = formule · VERT = lien inter-onglets", {}),
    ("", {}),
    ("Sources de marché : JLL Research, Investment I&L France Q1 2026 ; JLL Research, Investment Greater Paris Region Q1 2026.", {}),
    (DATA["DISCLAIMER"], dict(italic=True)),
]
for i, (txt, kw) in enumerate(rows, start=1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = Font(name="Arial", size=kw.get("size", 10), bold=kw.get("bold", False), italic=kw.get("italic", False),
                  color=NAVY if kw.get("bold") else "000000")
    c.alignment = Alignment(wrap_text=True, vertical="top")

# ============ Onglet 2 — Grille actifs ============
ws = wb.create_sheet("Grille actifs")
ws.sheet_properties.tabColor = GOLD
heads = ["ID", "Actif", "Type", "Profil", "Localisation", "Occupation", "WAULT (ans)", "Index",
         "Valeur (€m)", "Cap rate", "NOI (€m)", "Dette (€m)", "LTV actif", "NOI ÷ cap (€m)", "Écart check (€m)"]
ws.append(heads)
style_header(ws, 1, len(heads))
for i, a in enumerate(DATA["ASSETS"]):
    r = i + 2
    vals = [a["id"], a["name"], a["type"], a["profile"], a["location"], a["occupancy"],
            a["wault"] if a["wault"] is not None else "n.a.", a["leaseIndex"],
            a["value"], a["cap"], a["noi"], a["debt"]]
    for c, v in enumerate(vals, start=1):
        put(ws, f"{get_column_letter(c)}{r}", v,
            fmt={6: "0%", 7: "0.0", 9: M2, 10: PCT2, 11: M2, 12: M2}.get(c),
            color=BLUE_IN if c in (6, 7, 9, 10, 11, 12) else None)
    put(ws, f"M{r}", f"=L{r}/I{r}", PCT1)
    put(ws, f"N{r}", f"=K{r}/J{r}", M2)
    put(ws, f"O{r}", f"=I{r}-N{r}", M2)
put(ws, "B8", "Portefeuille (6 actifs)", bold=True)
put(ws, "I8", "=SUM(I2:I7)", M2, bold=True)
put(ws, "J8", "=K8/I8", PCT2, bold=True)
put(ws, "K8", "=SUM(K2:K7)", M2, bold=True)
put(ws, "L8", "=SUM(L2:L7)", M2, bold=True)
put(ws, "M8", "=L8/I8", PCT1, bold=True)
put(ws, "A10", "Convention : la Valeur est la donnée d'expertise (suivi en appui du valorisateur indépendant) ; "
               "NOI ÷ cap est le contrôle de cohérence (tolérance €0,2m, cf. onglet Contrôles).", italic=True, border=False)
for col, w in zip("ABCDEFGHIJKLMNO", [5, 22, 11, 15, 22, 11, 11, 7, 11, 9, 10, 10, 9, 13, 13]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "C2"

# ============ Onglet 3 — Dette & couverture ============
ws = wb.create_sheet("Dette & couverture")
ws.sheet_properties.tabColor = GOLD
heads = ["ID", "Taux all-in", "Part variable", "Couverture", "Expo nette (€m)", "Service (€m)",
         "Capex / TI (€m)", "BP NOI (€m)", "Écart NOI vs BP"]
ws.append(heads)
style_header(ws, 1, len(heads))
for i, a in enumerate(DATA["ASSETS"]):
    r = i + 2
    put(ws, f"A{r}", a["id"])
    put(ws, f"B{r}", a["allInRate"], PCT2, BLUE_IN)
    put(ws, f"C{r}", a["floatShare"], "0%", BLUE_IN)
    put(ws, f"D{r}", a["hedgeRatio"], "0%", BLUE_IN)
    put(ws, f"E{r}", f"='Grille actifs'!L{r}*C{r}*(1-D{r})", M2, GREEN_LINK)
    put(ws, f"F{r}", f"='Grille actifs'!L{r}*B{r}", M2, GREEN_LINK)
    put(ws, f"G{r}", a["capexReserve"], M2, BLUE_IN)
    put(ws, f"H{r}", a["bpNoi"], M2, BLUE_IN)
    put(ws, f"I{r}", f"='Grille actifs'!K{r}/H{r}-1", PCT1, GREEN_LINK)
put(ws, "A8", "Total", bold=True)
for col in "EFGH":
    put(ws, f"{col}8", f"=SUM({col}2:{col}7)", M2, bold=True)
put(ws, "I8", "='Grille actifs'!K8/H8-1", PCT1, GREEN_LINK, bold=True)
put(ws, "A10", "Expo nette = dette × part variable × (1 − couverture) : seule assiette sensible à la hausse des taux. "
               "Service = dette × taux all-in (interest-only).", italic=True, border=False)
for col, w in zip("ABCDEFGHI", [5, 11, 12, 11, 14, 12, 13, 12, 14]):
    ws.column_dimensions[col].width = w

# ============ Onglet 4 — Agrégats & covenants ============
ws = wb.create_sheet("Agrégats & covenants")
ws.sheet_properties.tabColor = NAVY
fund = DATA["FUND"]
def lab(r, label, value, fmt=None, color=None, bold=False):
    put(ws, f"A{r}", label, border=False, bold=bold)
    put(ws, f"B{r}", value, fmt, color, bold=bold)
put(ws, "A1", "AGRÉGATS DU FONDS — aucun chiffre saisi, tout est dérivé des onglets sources", bold=True, border=False)
lab(3,  "GAV (€m)", "='Grille actifs'!I8", M2, GREEN_LINK)
lab(4,  "NOI total (€m)", "='Grille actifs'!K8", M2, GREEN_LINK)
lab(5,  "Cap rate blended", "=B4/B3", PCT2)
lab(6,  "Dette totale (€m)", "='Grille actifs'!L8", M2, GREEN_LINK)
lab(7,  "LTV consolidée", "=B6/B3", PCT1, bold=True)
lab(8,  "Cash (€m)", fund["cash"], M2, BLUE_IN)
lab(9,  "Autres passifs (€m)", fund["otherLiabilities"], M2, BLUE_IN)
lab(10, "NAV (€m)", "=B3-B6+B8-B9", M2, bold=True)
lab(11, "Nombre de parts", fund["shares"], "#,##0", BLUE_IN)
lab(12, "VL par part (€)", "=B10*1000000/B11", '€#,##0.00', bold=True)
lab(14, "Service de dette (€m)", "='Dette & couverture'!F8", M2, GREEN_LINK)
lab(15, "DSCR covenant (interest-only)", "=B4/B14", X2, bold=True)
lab(16, "Capex / TI total (€m)", "='Dette & couverture'!G8", M2, GREEN_LINK)
lab(17, "Cash-flow après capex (€m)", "=B4-B16", M2)
lab(18, "DSCR après capex / TI", "=B17/B14", X2)
lab(19, "Exposition nette taux (€m)", "='Dette & couverture'!E8", M2, GREEN_LINK)
put(ws, "A21", "COVENANTS & RATIOS DE STRUCTURE", bold=True, border=False)
lab(22, "Covenant LTV max", fund["covenants"]["ltvMax"], "0%", BLUE_IN)
lab(23, "Seuil d'alerte LTV", fund["covenants"]["ltvAmberFrom"], "0%", BLUE_IN)
lab(24, "Statut LTV", '=IF(B7>B22,"BREACH",IF(B7>=B23,"ALERTE","CONFORME"))', bold=True)
lab(25, "Covenant DSCR min", fund["covenants"]["dscrMin"], X2, BLUE_IN)
lab(26, "Statut DSCR", '=IF(B15<B25,"BREACH","CONFORME")', bold=True)
lab(27, "Ratio d'emprise max (1er actif / GAV)", "=MAX('Grille actifs'!I2:I7)/B3", PCT1)
lab(28, "Limite interne d'emprise", fund["ratios"]["empriseMax"], "0%", BLUE_IN)
lab(29, "Statut emprise", '=IF(B27>B28,"DÉPASSEMENT","CONFORME")')
lab(30, "Levier AIFM simplifié (GAV / NAV)", "=B3/B10", X2)
lab(31, "Plafond prospectus", fund["ratios"]["aifmMax"], X2, BLUE_IN)
lab(32, "Statut levier", '=IF(B30>B31,"DÉPASSEMENT","CONFORME")')
lab(33, "LTV nette de trésorerie", "=(B6-B8)/B3", PCT1)
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 16

# ============ Onglet 5 — Scénario Bear ============
ws = wb.create_sheet("Scénario Bear")
ws.sheet_properties.tabColor = "C0392B"
bear = DATA["BEAR"]
put(ws, "A1", "SCÉNARIO BEAR — hypothèses modifiables (bleu), tout le reste recalcule", bold=True, border=False)
lab_rows = [("Choc sur les loyers", bear["rent"]), ("Vacance additionnelle", bear["vac"]),
            ("Expansion des cap rates", bear["cap"]), ("Hausse des taux courts", bear["rate"])]
for i, (l, v) in enumerate(lab_rows, start=2):
    put(ws, f"A{i}", l, border=False)
    put(ws, f"B{i}", v, PCT2, BLUE_IN)
heads = ["ID", "NOI stressé (€m)", "Cap stressé", "Valeur stressée (€m)", "Dette (€m)", "LTV actif"]
for c, h in enumerate(heads, start=1):
    ws.cell(row=8, column=c, value=h)
style_header(ws, 8, len(heads))
for i, a in enumerate(DATA["ASSETS"]):
    r = i + 9
    g = i + 2  # ligne correspondante dans Grille actifs
    put(ws, f"A{r}", a["id"])
    put(ws, f"B{r}", f"='Grille actifs'!K{g}*(1+$B$2)*(1-$B$3)", M2, GREEN_LINK)
    put(ws, f"C{r}", f"='Grille actifs'!J{g}+$B$4", PCT2, GREEN_LINK)
    put(ws, f"D{r}", f"=B{r}/C{r}", M2)
    put(ws, f"E{r}", f"='Grille actifs'!L{g}", M2, GREEN_LINK)
    put(ws, f"F{r}", f"=E{r}/D{r}", PCT1)
put(ws, "A15", "Total", bold=True)
put(ws, "B15", "=SUM(B9:B14)", M2, bold=True)
put(ws, "D15", "=SUM(D9:D14)", M2, bold=True)
put(ws, "E15", "=SUM(E9:E14)", M2, bold=True)
put(ws, "F15", "=E15/D15", PCT1, bold=True)
lab2 = [
    (17, "Service de dette stressé (€m)", "='Dette & couverture'!F8+$B$5*'Dette & couverture'!E8", M2),
    (18, "DSCR covenant (IO)", "=B15/B17", X2),
    (19, "DSCR après capex / TI", "=(B15-'Dette & couverture'!G8)/B17", X2),
    (20, "NAV stressée (€m)", "=D15-E15+'Agrégats & covenants'!B8-'Agrégats & covenants'!B9", M2),
    (21, "VL stressée (€)", "=B20*1000000/'Agrégats & covenants'!B11", '€#,##0.00'),
    (22, "Statut covenant LTV", "=IF(F15>'Agrégats & covenants'!B22,\"BREACH\",IF(F15>='Agrégats & covenants'!B23,\"ALERTE\",\"CONFORME\"))", None),
    (23, "Statut covenant DSCR", "=IF(B18<'Agrégats & covenants'!B25,\"BREACH\",\"CONFORME\")", None),
]
for r, l, f_, fmt in lab2:
    put(ws, f"A{r}", l, border=False)
    put(ws, f"B{r}", f_, fmt, bold=r in (18, 19, 22))
put(ws, "A25", "Lecture : la LTV casse (canal valeur), le DSCR tient (dette interest-only + couverture de taux). "
               "Les deux DSCR dérivent du même NOI stressé.", italic=True, border=False)
ws.column_dimensions["A"].width = 30
for col, w in zip("BCDEF", [17, 12, 18, 12, 10]):
    ws.column_dimensions[col].width = w

# ============ Onglet 6 — Contrôles ============
ws = wb.create_sheet("Contrôles")
ws.sheet_properties.tabColor = "2E7D5B"
heads = ["Contrôle", "Valeur du classeur", "Référence cockpit", "Tolérance", "Statut"]
ws.append(heads)
style_header(ws, 1, len(heads))
checks = [
    ("GAV (€m)", "='Agrégats & covenants'!B3", 317.4, 0.05, M2),
    ("NOI total (€m)", "='Agrégats & covenants'!B4", 17.18, 0.005, M2),
    ("Dette totale (€m)", "='Agrégats & covenants'!B6", 149.0, 0.001, M2),
    ("LTV consolidée", "='Agrégats & covenants'!B7", 0.46944, 0.0005, PCT2),
    ("NAV (€m)", "='Agrégats & covenants'!B10", 170.5, 0.05, M2),
    ("VL par part (€)", "='Agrégats & covenants'!B12", 100.0, 0.01, '€#,##0.00'),
    ("DSCR covenant (IO)", "='Agrégats & covenants'!B15", 2.5887, 0.005, X2),
    ("DSCR après capex / TI", "='Agrégats & covenants'!B18", 2.2346, 0.005, X2),
    ("Exposition nette taux (€m)", "='Agrégats & covenants'!B19", 36.95, 0.01, M2),
    ("Écart max valeur vs NOI÷cap (€m)", "=MAX(MAX('Grille actifs'!O2:O7),-MIN('Grille actifs'!O2:O7))", 0.0, 0.2, M2),
    ("Bear — GAV stressée (€m)", "='Scénario Bear'!D15", 221.28, 0.05, M2),
    ("Bear — LTV consolidée", "='Scénario Bear'!F15", 0.67335, 0.001, PCT2),
    ("Bear — DSCR covenant", "='Scénario Bear'!B18", 2.0642, 0.005, X2),
    ("Bear — DSCR après capex", "='Scénario Bear'!B19", 1.7375, 0.005, X2),
]
for i, (label, formula, ref, tol, fmt) in enumerate(checks):
    r = i + 2
    put(ws, f"A{r}", label)
    put(ws, f"B{r}", formula, fmt, GREEN_LINK)
    put(ws, f"C{r}", ref, fmt, BLUE_IN)
    put(ws, f"D{r}", tol, "0.000", BLUE_IN)
    put(ws, f"E{r}", f'=IF(ABS(B{r}-C{r})<=D{r},"PASS","FAIL")', bold=True)
put(ws, "A17", "Synthèse", bold=True)
put(ws, "E17", '=IF(COUNTIF(E2:E15,"FAIL")=0,"✓ TOUS LES CONTRÔLES PASSENT",COUNTIF(E2:E15,"FAIL")&" ÉCHEC(S)")', bold=True)
put(ws, "A19", "Références = valeurs de la recette automatique du cockpit (12/12 tests). "
               "Si un input change, les statuts basculent en FAIL tant que la référence n'est pas réalignée.", italic=True, border=False)
for col, w in zip("ABCDE", [34, 18, 18, 11, 28]):
    ws.column_dimensions[col].width = w

wb.calculation.fullCalcOnLoad = True  # pas de valeurs en cache : Excel recalcule à l'ouverture
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("OK:", os.path.abspath(OUT))
